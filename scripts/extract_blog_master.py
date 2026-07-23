#!/usr/bin/env python3
"""Extrae BLOG_MASTER / FAQ_BLOG / TAGS_BLOG de MASTER_GLOBAL_BLOG_ECOFARMA.xlsx
a un JSON intermedio que import-blog-master.ts consume.

Existe como paso separado porque exceljs (Node) no logra leer este archivo
especifico (falla en workbook.xml -- ver notas de la implementacion), pero
openpyxl si lo lee sin problema. Uso interno, invocado automaticamente por
import-blog-master.ts -- no se corre a mano normalmente.

Uso: python3 extract_blog_master.py <ruta.xlsx> <limit> <ruta_salida.json>
"""
import json
import sys

import openpyxl

FIRST_DATA_ROW = 4  # fila 1 = titulo, fila 2 = vacia, fila 3 = encabezados


def cell(row, idx):
    """idx es 1-based (columna A = 1), como en el Excel."""
    value = row[idx - 1]
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def cell_int(row, idx):
    s = cell(row, idx)
    if s is None:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def split_list(value, sep):
    if not value:
        return []
    return [v.strip() for v in value.split(sep) if v.strip()]


def main():
    xlsx_path, limit_str, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    limit = int(limit_str)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    master = wb["BLOG_MASTER"]
    faq = wb["FAQ_BLOG"]
    tags = wb["TAGS_BLOG"]

    posts = []
    global_ids = set()
    for r in master.iter_rows(min_row=FIRST_DATA_ROW, max_row=FIRST_DATA_ROW + limit - 1, values_only=True):
        global_id = cell(r, 1)
        if not global_id:
            break
        global_ids.add(global_id)
        posts.append(
            {
                "globalId": global_id,
                "sourceContentId": cell(r, 3),
                "hub": cell(r, 5) or "SIN HUB",
                "subHub": cell(r, 6),
                "title": cell(r, 7) or global_id,
                "slug": cell(r, 8) or global_id.lower(),
                "headings": split_list(cell(r, 9), "\n"),
                "cluster": cell(r, 11),
                "tagPrincipal": cell(r, 12),
                "tagsSecondary": split_list(cell(r, 13), ";"),
                "seoPriority": cell(r, 14),
                "regulatoryLevel": cell(r, 15),
                "intention": cell(r, 16),
                "productPolicy": cell(r, 24),
                "validationRequired": cell(r, 25),
                "aiGenerationRule": cell(r, 26),
                "notes": cell(r, 27),
                "sourceFile": cell(r, 28),
                "sourceRow": cell_int(r, 29),
            }
        )

    faqs = []
    for r in faq.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        global_id = cell(r, 1)
        if not global_id:
            break
        if global_id not in global_ids:
            continue
        faqs.append(
            {
                "globalId": global_id,
                "faqBlockId": cell(r, 7),
                "questionNumber": cell_int(r, 8),
                "question": cell(r, 9) or "",
                "sourceQuestion": cell(r, 11),
                "sourceReferences": cell(r, 12),
                "regulatoryLevel": cell(r, 14),
                "productPolicy": cell(r, 15),
                "validationRequired": cell(r, 16),
                "notes": cell(r, 17),
                "sourceFile": cell(r, 18),
                "sourceRow": cell_int(r, 19),
            }
        )

    tag_rows = []
    for r in tags.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        tag = cell(r, 1)
        if not tag:
            break
        tag_rows.append(
            {
                "tag": tag,
                "role": cell(r, 3),
                "linkedHubs": cell(r, 4),
                "linkedSubHubs": cell(r, 5),
                "contentCount": cell_int(r, 6),
                "examples": cell(r, 7),
                "usageRule": cell(r, 8),
            }
        )

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"posts": posts, "faqs": faqs, "tags": tag_rows}, f, ensure_ascii=False)

    print(
        f"Extraidos: {len(posts)} posts, {len(faqs)} faqs, {len(tag_rows)} tags -> {out_path}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
