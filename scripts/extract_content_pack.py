#!/usr/bin/env python3
"""Extrae HUBS / BLOGS / ENCICLOPEDIA / TAGS de un paquete de contenido tipo
ENTREGA_TABLAS_SEPARADAS.xlsx a un JSON intermedio que import-content-pack.ts
consume.

A diferencia de extract_blog_master.py (columnas por indice fijo, datos desde
la fila 4), este paquete tiene headers en la fila 1 y datos desde la fila 2 --
se lee por NOMBRE de columna. El cuerpo de cada pagina/articulo llega en un
solo bloque de texto con `## ` (H2) y, dentro de la seccion de preguntas
frecuentes, `### ` (pregunta) -- se parsea aqui en secciones y FAQs (a
diferencia del Excel maestro viejo, las respuestas de las FAQs ya vienen
escritas).

content_id no es unico globalmente: algunos temas tienen una version V12 y
una V13 (el mismo id repetido en dos source_file distintos, nunca dentro del
mismo archivo) -- se desambigua con un sufijo `__dupN` determinista por orden
de aparicion, no se descarta ninguna fila.

Uso: python3 extract_content_pack.py <ruta.xlsx> <limit_por_hoja> <ruta_salida.json>
"""
import json
import re
import sys
import unicodedata

import openpyxl

SHEET_CONTENT_TYPES = {
    "HUBS": "HUB",
    "BLOGS": "BLOG",
    "ENCICLOPEDIA": "ENCICLOPEDIA",
}


def norm(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def is_faq_heading(heading):
    return "pregunta" in strip_accents(heading).lower()


def split_list(value, sep):
    if not value:
        return []
    return [v.strip() for v in value.split(sep) if v.strip()]


def parse_body(body):
    """Devuelve (sections, faqs) a partir del blob con `## `/`### ` inline."""
    sections = []
    faqs = []
    if not body:
        return sections, faqs

    parts = re.split(r"^## (.+)$", body, flags=re.MULTILINE)
    # parts[0] = texto antes del primer H2 (intro suelta, se descarta -- las
    # hojas HUBS ya traen su propio campo intro_page_a_remplir separado).
    for i in range(1, len(parts), 2):
        heading = parts[i].strip()
        section_body = parts[i + 1].strip() if i + 1 < len(parts) else ""

        if is_faq_heading(heading):
            qa_parts = re.split(r"^### (.+)$", section_body, flags=re.MULTILINE)
            for j in range(1, len(qa_parts), 2):
                question = qa_parts[j].strip()
                answer = qa_parts[j + 1].strip() if j + 1 < len(qa_parts) else ""
                if question:
                    faqs.append({"question": question, "answer": answer or None})
            continue

        sections.append({"heading": heading, "body": section_body or None})

    return sections, faqs


def read_sheet(ws, content_type_base, limit, seen_ids):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    def col(row, name):
        i = idx.get(name)
        return norm(row[i]) if i is not None else None

    posts = []
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        content_id = col(row, "content_id")
        if not content_id:
            continue
        if limit is not None and count >= limit:
            break
        count += 1

        global_id = content_id
        if global_id in seen_ids:
            seen_ids[global_id] += 1
            global_id = f"{content_id}__dup{seen_ids[content_id]}"
        else:
            seen_ids[global_id] = 1

        body_col = "corps_page_a_remplir" if "corps_page_a_remplir" in idx else "corps_article_a_remplir"
        sections, faqs = parse_body(col(row, body_col))
        internal_links_notes = " | ".join(
            filter(
                None,
                [
                    col(row, "liens_internes_a_prevoir"),
                    col(row, "articles_blog_lies"),
                    col(row, "fiches_encyclopedie_liees"),
                ],
            )
        ) or None

        posts.append(
            {
                "globalId": global_id,
                "sourceContentId": content_id,
                "sourceFile": col(row, "source_file"),
                "contentType": col(row, "type_contenu") or content_type_base,
                "hub": col(row, "hub_principal") or "SIN HUB",
                "subHub": col(row, "sous_hub"),
                "title": col(row, "titre_h1_seo_optimise_a_remplir") or col(row, "titre_h1") or content_id,
                "slug": col(row, "slug_a_remplir"),
                "tagPrincipal": col(row, "tag_principal"),
                "tagsSecondary": split_list(col(row, "tags_secondaires"), ";"),
                "intro": col(row, "intro_page_a_remplir"),
                "metaTitle": col(row, "meta_title_a_remplir"),
                "metaDescription": col(row, "meta_description_a_remplir"),
                "internalLinksNotes": internal_links_notes,
                "sourcesConsultadas": col(row, "sources_finales_a_citer"),
                "regulatoryLevel": col(row, "niveau_reglementaire"),
                "productPolicy": col(row, "politique_produits"),
                "sections": sections,
                "faqs": faqs,
            }
        )

    return posts


def read_tags(ws, limit):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    def col(row, name):
        i = idx.get(name)
        return norm(row[i]) if i is not None else None

    tags = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tag = col(row, "tag_principal")
        if not tag:
            continue
        if limit is not None and len(tags) >= limit:
            break
        tags.append(
            {
                "tag": tag,
                "role": col(row, "type_tag"),
                "linkedHubs": col(row, "hub_ou_sous_hub_associe"),
                "examples": col(row, "tags_secondaires_possibles"),
                "usageRule": col(row, "usage"),
            }
        )
    return tags


def main():
    xlsx_path, limit_str, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    limit = None if limit_str in ("", "0", "none", "None") else int(limit_str)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    seen_ids = {}
    posts = []
    for sheet_name, content_type_base in SHEET_CONTENT_TYPES.items():
        posts.extend(read_sheet(wb[sheet_name], content_type_base, limit, seen_ids))

    tags = read_tags(wb["TAGS"], limit)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"posts": posts, "tags": tags}, f, ensure_ascii=False)

    n_sections = sum(len(p["sections"]) for p in posts)
    n_faqs = sum(len(p["faqs"]) for p in posts)
    print(
        f"Extraidos: {len(posts)} posts ({n_sections} secciones, {n_faqs} faqs), {len(tags)} tags -> {out_path}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
